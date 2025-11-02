from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class ExcelReportGenerator:
    """Generador de reportes en Excel"""
    
    def __init__(self, title="Reporte"):
        self.title = title
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remover hoja por defecto
        
        # Estilos
        self.header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        self.title_font = Font(name='Arial', size=16, bold=True)
        self.title_alignment = Alignment(horizontal='center', vertical='center')
        
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_sheet(self, name):
        """Crear una nueva hoja"""
        return self.workbook.create_sheet(title=name)
    
    def add_title_row(self, sheet, title, row=1):
        """Agregar título en la primera fila"""
        sheet.merge_cells(f'A{row}:F{row}')
        cell = sheet[f'A{row}']
        cell.value = title
        cell.font = self.title_font
        cell.alignment = self.title_alignment
    
    def add_metadata(self, sheet, tenant_name, generated_by, row=2):
        """Agregar metadata"""
        sheet[f'A{row}'] = 'Organización:'
        sheet[f'B{row}'] = tenant_name
        sheet[f'A{row+1}'] = 'Generado por:'
        sheet[f'B{row+1}'] = generated_by
        sheet[f'A{row+2}'] = 'Fecha:'
        sheet[f'B{row+2}'] = datetime.now().strftime('%d/%m/%Y %H:%M')
    
    def add_table(self, sheet, data, start_row=6, headers=None):
        """Agregar tabla con datos"""
        if not data:
            return
        
        # Agregar headers
        if headers:
            for col_idx, header in enumerate(headers, 1):
                cell = sheet.cell(row=start_row, column=col_idx)
                cell.value = header
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.header_alignment
                cell.border = self.border
        
        # Agregar datos
        for row_idx, row_data in enumerate(data, start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = self.border
                
                # Alternar colores de fondo
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
        
        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1 if headers else len(data[0]) + 1):
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = 20
    
    def generate(self):
        """Generar el Excel y retornar el buffer"""
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer
    
    def save(self, filename):
        """Guardar el Excel en un archivo"""
        self.workbook.save(filename)
        return filename


def generate_documents_excel(data, tenant_name, user_name):
    """Generar reporte de documentos en Excel"""
    excel = ExcelReportGenerator(title="Reporte de Documentos")
    
    # Hoja 1: Resumen
    sheet1 = excel.create_sheet("Resumen")
    excel.add_title_row(sheet1, "Reporte de Documentos Clínicos")
    excel.add_metadata(sheet1, tenant_name, user_name)
    
    # Estadísticas
    sheet1['A6'] = 'Total de Documentos:'
    sheet1['B6'] = data.get('total', 0)
    sheet1['B6'].font = Font(bold=True, size=14)
    
    # Hoja 2: Documentos por Tipo
    if data.get('by_type'):
        sheet2 = excel.create_sheet("Por Tipo")
        excel.add_title_row(sheet2, "Documentos por Tipo")
        
        table_data = [[item['document_type'], item['count']] for item in data['by_type']]
        excel.add_table(sheet2, table_data, headers=['Tipo de Documento', 'Cantidad'])
    
    # Hoja 3: Documentos Recientes
    if data.get('recent_documents'):
        sheet3 = excel.create_sheet("Recientes")
        excel.add_title_row(sheet3, "Documentos Recientes")
        
        table_data = [
            [
                doc.get('document_date', '')[:10],
                doc.get('document_type', ''),
                doc.get('patient_name', ''),
                doc.get('doctor_name', ''),
                doc.get('specialty', '')
            ]
            for doc in data['recent_documents']
        ]
        excel.add_table(sheet3, table_data, 
                       headers=['Fecha', 'Tipo', 'Paciente', 'Doctor', 'Especialidad'])
    
    return excel.generate()
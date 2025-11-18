"""
Servicio de exportación de reportes a PDF y Excel
"""
import io
from datetime import datetime
from typing import List, Dict, Any
from django.http import HttpResponse

# PDF
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# Excel
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


class ReportExportService:
    """Servicio para exportar reportes a diferentes formatos"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Configurar estilos personalizados para PDF"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2d6147'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        ))

        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2d6147'),
            spaceAfter=12,
            fontName='Helvetica-Bold'
        ))

    def export_to_pdf(
        self,
        data: List[Dict[str, Any]],
        title: str,
        columns: List[str],
        metadata: Dict[str, Any] = None
    ) -> bytes:
        """
        Exportar datos a PDF

        Args:
            data: Lista de diccionarios con los datos
            title: Título del reporte
            columns: Lista de columnas a incluir
            metadata: Metadatos adicionales (fecha, usuario, etc.)

        Returns:
            bytes: Contenido del PDF
        """
        buffer = io.BytesIO()

        # Función para el footer con paginación
        def add_page_number(canvas, doc):
            """Agrega número de página al footer"""
            page_num = canvas.getPageNumber()
            text = f"Página {page_num}"
            canvas.saveState()
            canvas.setFont('Helvetica', 9)
            canvas.drawRightString(letter[0] - 50, 30, text)
            # Agregar fecha de generación
            now = datetime.now()
            footer_text = f"Generado: {now.strftime('%d/%m/%Y %H:%M')}"
            canvas.drawString(50, 30, footer_text)
            canvas.restoreState()

        doc = SimpleDocTemplate(buffer, pagesize=letter)
        story = []

        # Tenant name (grande)
        tenant_name = metadata.get('tenant_name', 'Sistema de Historias Clínicas') if metadata else 'Sistema de Historias Clínicas'
        tenant_style = ParagraphStyle(
            name='TenantName',
            fontSize=20,
            textColor=colors.black,
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph(tenant_name, tenant_style))
        story.append(Spacer(1, 0.2 * inch))

        # Título del reporte
        report_title_style = ParagraphStyle(
            name='ReportTitle',
            fontSize=16,
            textColor=colors.HexColor('#444444'),
            spaceAfter=20,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        story.append(Paragraph(title, report_title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Información del generador
        if metadata:
            now = datetime.now()
            generator_info = f"""
            <b>Generado por:</b> {metadata.get('user_name', 'N/A')}<br/>
            <b>Fecha:</b> {now.strftime('%d de %B de %Y')}<br/>
            <b>Hora:</b> {now.strftime('%H:%M:%S')}<br/>
            <b>Total de registros:</b> {len(data)}
            """
            story.append(Paragraph(generator_info, self.styles['Normal']))
            story.append(Spacer(1, 0.3 * inch))

        # Tabla de datos
        if data:
            # Encabezados
            table_data = [[col.replace('_', ' ').title() for col in columns]]

            # Filas
            for row in data:
                table_data.append([str(row.get(col, '')) for col in columns])

            # Crear tabla
            table = Table(table_data)
            table.setStyle(TableStyle([
                # Encabezado - GRIS en lugar de azul
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#d1d5db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TOPPADDING', (0, 0), (-1, 0), 12),

                # Contenido
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ]))

            story.append(table)
        else:
            story.append(Paragraph("No hay datos para mostrar", self.styles['Normal']))

        # Construir PDF con paginación
        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        pdf = buffer.getvalue()
        buffer.close()

        return pdf

    def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        title: str,
        columns: List[str],
        metadata: Dict[str, Any] = None
    ) -> bytes:
        """
        Exportar datos a Excel

        Args:
            data: Lista de diccionarios con los datos
            title: Título del reporte
            columns: Lista de columnas a incluir
            metadata: Metadatos adicionales

        Returns:
            bytes: Contenido del archivo Excel
        """
        # Convertir a DataFrame
        if data:
            df = pd.DataFrame(data)[columns]
            # Limpiar nombres de columnas
            df.columns = [col.replace('_', ' ').title() for col in df.columns]
        else:
            df = pd.DataFrame()

        # Crear workbook
        buffer = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limita a 31 caracteres

        # Estilos
        header_fill = PatternFill(start_color='2d6147', end_color='2d6147', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')

        row_num = 1

        # Título
        ws.merge_cells(f'A{row_num}:E{row_num}')
        title_cell = ws[f'A{row_num}']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=16, color='2d6147')
        title_cell.alignment = center_alignment
        row_num += 2

        # Metadata
        if metadata:
            ws[f'A{row_num}'] = 'Generado:'
            ws[f'B{row_num}'] = metadata.get('generated_at', datetime.now().strftime('%Y-%m-%d %H:%M'))
            row_num += 1

            ws[f'A{row_num}'] = 'Usuario:'
            ws[f'B{row_num}'] = metadata.get('user_name', 'N/A')
            row_num += 1

            ws[f'A{row_num}'] = 'Registros:'
            ws[f'B{row_num}'] = len(data)
            row_num += 2

        # Datos
        if not df.empty:
            # Encabezados
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

            row_num += 1

            # Filas de datos
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), row_num):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.value = value

                    # Alternar colores de fila
                    if (r_idx - row_num) % 2 == 0:
                        cell.fill = PatternFill(start_color='f9fafb', end_color='f9fafb', fill_type='solid')

            # Ajustar ancho de columnas
            for col_idx, column_cells in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for cell in column_cells:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        else:
            ws[f'A{row_num}'] = 'No hay datos para mostrar'

        # Guardar
        wb.save(buffer)
        excel = buffer.getvalue()
        buffer.close()

        return excel

    def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        columns: List[str]
    ) -> str:
        """
        Exportar datos a CSV

        Args:
            data: Lista de diccionarios con los datos
            columns: Lista de columnas a incluir

        Returns:
            str: Contenido del CSV
        """
        if not data:
            return ''

        df = pd.DataFrame(data)[columns]
        return df.to_csv(index=False)

    def create_http_response(
        self,
        content: bytes,
        filename: str,
        content_type: str
    ) -> HttpResponse:
        """
        Crear HttpResponse para descarga

        Args:
            content: Contenido del archivo
            filename: Nombre del archivo
            content_type: Tipo MIME

        Returns:
            HttpResponse configurado para descarga
        """
        response = HttpResponse(content, content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
